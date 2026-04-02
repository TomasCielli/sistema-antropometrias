import io
import json
import os
import tempfile
from datetime import date
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session
from openpyxl import Workbook, load_workbook
from database import get_connection
from models.antropometrias import CAMPOS_NUMERICOS, LABELS

datos_bp = Blueprint("datos", __name__)

# Jugador fields to compare for conflict detection
CAMPOS_JUGADOR = ["nombre", "apellido", "sexo", "posicion_actual",
                  "categoria_actual", "fecha_nacimiento", "objetivo",
                  "telefono", "observaciones"]

LABELS_JUGADOR = {
    "nombre": "Nombre",
    "apellido": "Apellido",
    "sexo": "Sexo",
    "posicion_actual": "Posición actual",
    "categoria_actual": "Categoría actual",
    "fecha_nacimiento": "Fecha de nacimiento",
    "objetivo": "Objetivo",
    "telefono": "Teléfono",
    "observaciones": "Observaciones",
}


def _normalizar(val):
    """Normalize a value for comparison: None and empty string are equal."""
    if val is None:
        return ""
    return str(val).strip()


@datos_bp.route("/datos/exportar")
def exportar():
    conn = get_connection()
    cursor = conn.cursor()

    wb = Workbook()

    # Sheet 1: Jugadores
    ws_jug = wb.active
    ws_jug.title = "Jugadores"
    jug_headers = ["id", "apellido", "nombre", "dni", "sexo",
                   "posicion_actual", "categoria_actual",
                   "fecha_nacimiento", "objetivo", "telefono",
                   "observaciones", "creado_en"]
    ws_jug.append(jug_headers)

    cursor.execute("SELECT * FROM jugadores ORDER BY apellido, nombre")
    for row in cursor.fetchall():
        ws_jug.append([row[h] for h in jug_headers])

    # Sheet 2: Antropometrias
    ws_ant = wb.create_sheet("Antropometrias")
    ant_headers = ["id", "jugador_id", "jugador_apellido", "jugador_nombre",
                   "fecha", "posicion", "categoria"] + CAMPOS_NUMERICOS
    ws_ant.append(ant_headers)

    cursor.execute("""
        SELECT a.*, j.apellido AS jugador_apellido, j.nombre AS jugador_nombre
        FROM antropometrias a
        JOIN jugadores j ON a.jugador_id = j.id
        ORDER BY a.jugador_id, a.fecha DESC
    """)
    for row in cursor.fetchall():
        fila = [
            row["id"], row["jugador_id"],
            row["jugador_apellido"], row["jugador_nombre"],
            row["fecha"], row["posicion"], row["categoria"]
        ]
        for campo in CAMPOS_NUMERICOS:
            fila.append(row[campo])
        ws_ant.append(fila)

    conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"albatros_antropometrias_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@datos_bp.route("/datos/importar", methods=["GET"])
def importar():
    # Clean up any leftover temp file from a previous import
    temp_path = session.pop("import_temp", None)
    if temp_path and os.path.exists(temp_path):
        os.remove(temp_path)
    return render_template("datos/importar.html")


@datos_bp.route("/datos/importar", methods=["POST"])
def procesar_importacion():
    file = request.files.get("archivo")

    if not file or file.filename == "":
        flash("No se seleccionó ningún archivo", "danger")
        return render_template("datos/importar.html")

    if not file.filename.endswith(".xlsx"):
        flash("El archivo debe ser .xlsx", "danger")
        return render_template("datos/importar.html")

    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        flash("El archivo no es un Excel válido", "danger")
        return render_template("datos/importar.html")

    hojas_faltantes = []
    if "Jugadores" not in wb.sheetnames:
        hojas_faltantes.append("Jugadores")
    if "Antropometrias" not in wb.sheetnames:
        hojas_faltantes.append("Antropometrias")

    if hojas_faltantes:
        flash(f"Faltan las hojas: {', '.join(hojas_faltantes)}", "danger")
        return render_template("datos/importar.html")

    conn = get_connection()
    cursor = conn.cursor()

    # --- Parse Jugadores sheet ---
    ws_jug = wb["Jugadores"]
    jug_headers = [cell.value for cell in ws_jug[1]]
    nuevos = []
    conflictos = []
    errores_jug = []

    for fila_num, row in enumerate(ws_jug.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        datos = dict(zip(jug_headers, row))
        dni = _normalizar(datos.get("dni"))
        nombre = _normalizar(datos.get("nombre"))
        apellido = _normalizar(datos.get("apellido"))
        sexo = _normalizar(datos.get("sexo"))
        fecha_nac = _normalizar(datos.get("fecha_nacimiento"))

        if not dni:
            errores_jug.append(f"Fila {fila_num}: campo \"DNI\" — campo obligatorio vacío")
            continue
        if not nombre:
            errores_jug.append(f"Fila {fila_num} (DNI: {dni}): campo \"Nombre\" — campo obligatorio vacío")
            continue
        if not apellido:
            errores_jug.append(f"Fila {fila_num} (DNI: {dni}): campo \"Apellido\" — campo obligatorio vacío")
            continue
        if not sexo:
            errores_jug.append(f"Fila {fila_num} (DNI: {dni}): campo \"Sexo\" — campo obligatorio vacío")
            continue
        if not fecha_nac:
            errores_jug.append(f"Fila {fila_num} (DNI: {dni}): campo \"Fecha de nacimiento\" — campo obligatorio vacío")
            continue

        jugador_importado = {
            "nombre": nombre,
            "apellido": apellido,
            "dni": dni,
            "sexo": sexo,
            "fecha_nacimiento": fecha_nac,
            "posicion_actual": _normalizar(datos.get("posicion_actual")) or None,
            "categoria_actual": _normalizar(datos.get("categoria_actual")) or None,
            "objetivo": _normalizar(datos.get("objetivo")) or None,
            "telefono": _normalizar(datos.get("telefono")) or None,
            "observaciones": _normalizar(datos.get("observaciones")) or None,
        }

        # Check if DNI exists in DB
        cursor.execute("SELECT * FROM jugadores WHERE dni = ?", (dni,))
        existente = cursor.fetchone()

        if not existente:
            nuevos.append(jugador_importado)
        else:
            existente_dict = dict(existente)
            # Find differing fields
            diferencias = []
            for campo in CAMPOS_JUGADOR:
                val_db = _normalizar(existente_dict.get(campo))
                val_imp = _normalizar(jugador_importado.get(campo))
                if val_db != val_imp:
                    diferencias.append({
                        "campo": campo,
                        "label": LABELS_JUGADOR.get(campo, campo),
                        "valor_db": existente_dict.get(campo) or "—",
                        "valor_importado": jugador_importado.get(campo) or "—",
                    })

            if diferencias:
                conflictos.append({
                    "dni": dni,
                    "nombre_db": f"{existente_dict['apellido']}, {existente_dict['nombre']}",
                    "nombre_importado": f"{apellido}, {nombre}",
                    "id_db": existente_dict["id"],
                    "diferencias": diferencias,
                    "datos_importados": jugador_importado,
                })
            # If identical, silently skip

    # --- Parse Antropometrias sheet ---
    ws_ant = wb["Antropometrias"]
    ant_headers = [cell.value for cell in ws_ant[1]]
    antropometrias_pendientes = []
    errores_ant = []

    for fila_num, row in enumerate(ws_ant.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        datos = dict(zip(ant_headers, row))
        jugador_id = datos.get("jugador_id")
        fecha = datos.get("fecha")

        if not jugador_id:
            errores_ant.append(f"Fila {fila_num}: campo \"jugador_id\" — campo obligatorio vacío")
            continue
        if not fecha:
            errores_ant.append(f"Fila {fila_num}: campo \"Fecha\" — campo obligatorio vacío")
            continue

        try:
            jugador_id = int(jugador_id)
        except (ValueError, TypeError):
            errores_ant.append(f"Fila {fila_num}: campo \"jugador_id\" — se esperaba un número, se recibió '{jugador_id}'")
            continue

        fecha_str = _normalizar(fecha)

        # Validate numeric fields
        campos_validos = {"jugador_id": jugador_id, "fecha": fecha_str}
        fila_valida = True
        for campo in ["posicion", "categoria"]:
            val = datos.get(campo)
            if val:
                campos_validos[campo] = str(val)

        for campo in CAMPOS_NUMERICOS:
            val = datos.get(campo)
            if val is not None and val != "" and val != "—":
                try:
                    val_float = float(val)
                    if val_float < 0:
                        label = LABELS.get(campo, (campo,))[0]
                        errores_ant.append(
                            f"Fila {fila_num}: campo \"{label}\" — el valor no puede ser negativo"
                        )
                        fila_valida = False
                        break
                    campos_validos[campo] = val_float
                except (ValueError, TypeError):
                    label = LABELS.get(campo, (campo,))[0]
                    errores_ant.append(
                        f"Fila {fila_num}: campo \"{label}\" — se esperaba un número, se recibió '{val}'"
                    )
                    fila_valida = False
                    break

        if fila_valida:
            antropometrias_pendientes.append(campos_validos)

    conn.close()

    # --- If conflicts exist → save to temp file, show conflict resolution page ---
    if conflictos:
        temp_data = {
            "nuevos": nuevos,
            "antropometrias": antropometrias_pendientes,
            "errores_jug": errores_jug,
            "errores_ant": errores_ant,
        }
        tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, encoding="utf-8")
        json.dump(temp_data, tmp, ensure_ascii=False)
        tmp.close()
        session["import_temp"] = tmp.name

        return render_template("datos/conflictos.html",
                               conflictos=conflictos,
                               errores_jug=errores_jug,
                               errores_ant=errores_ant,
                               total_nuevos=len(nuevos),
                               total_antro=len(antropometrias_pendientes))

    # --- No conflicts → process directly ---
    resultados = _procesar_datos(nuevos, {}, antropometrias_pendientes)
    _flash_resultado(resultados, errores_jug, errores_ant)
    return redirect(url_for("datos.importar"))


@datos_bp.route("/datos/resolver_conflictos", methods=["POST"])
def resolver_conflictos():
    temp_path = session.get("import_temp")
    if not temp_path or not os.path.exists(temp_path):
        flash("La sesión de importación expiró. Por favor, volvé a subir el archivo.", "warning")
        return redirect(url_for("datos.importar"))

    with open(temp_path, "r", encoding="utf-8") as f:
        temp_data = json.load(f)

    os.remove(temp_path)
    session.pop("import_temp", None)

    nuevos = temp_data["nuevos"]
    antropometrias_pendientes = temp_data["antropometrias"]
    errores_jug = temp_data["errores_jug"]
    errores_ant = temp_data["errores_ant"]

    # Build decisions dict: dni → "mantener" | "sobreescribir"
    decisiones = {}
    for key, val in request.form.items():
        if key.startswith("decision_"):
            dni = key[len("decision_"):]
            decisiones[dni] = val

    resultados = _procesar_datos(nuevos, decisiones, antropometrias_pendientes)
    _flash_resultado(resultados, errores_jug, errores_ant)
    return redirect(url_for("datos.importar"))


def _procesar_datos(nuevos, decisiones, antropometrias_pendientes):
    """Insert nuevos, apply decisions for conflicts, insert anthropometries."""
    conn = get_connection()
    cursor = conn.cursor()

    jug_insertados = 0
    jug_sobreescritos = 0
    jug_omitidos = 0

    # Insert new players
    for j in nuevos:
        try:
            cursor.execute("""
                INSERT INTO jugadores (nombre, apellido, dni, sexo, posicion_actual,
                                       categoria_actual, fecha_nacimiento, objetivo,
                                       telefono, observaciones)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (j["nombre"], j["apellido"], j["dni"], j["sexo"],
                  j.get("posicion_actual"), j.get("categoria_actual"),
                  j["fecha_nacimiento"], j.get("objetivo"),
                  j.get("telefono"), j.get("observaciones")))
            jug_insertados += 1
        except Exception:
            jug_omitidos += 1

    # Apply conflict decisions
    for dni, decision in decisiones.items():
        if decision == "sobreescribir":
            # We need the imported data — it was stored in the conflict form as hidden fields
            # The form sends: sobreescribir_<dni>_<campo> = valor
            cursor.execute("SELECT id FROM jugadores WHERE dni = ?", (dni,))
            row = cursor.fetchone()
            if row:
                jugador_id = row["id"]
                # Get imported values from form hidden fields
                campos_update = {}
                for campo in CAMPOS_JUGADOR:
                    key = f"datos_{dni}_{campo}"
                    val = request.form.get(key)
                    if val is not None:
                        campos_update[campo] = val.strip() or None

                if campos_update:
                    sets = ", ".join(f"{c} = ?" for c in campos_update)
                    vals = list(campos_update.values()) + [jugador_id]
                    cursor.execute(f"UPDATE jugadores SET {sets} WHERE id = ?", vals)
                    jug_sobreescritos += 1
        else:
            jug_omitidos += 1

    conn.commit()

    # Process anthropometries
    ant_insertadas = 0
    ant_omitidas = 0
    ant_errores = 0

    for a in antropometrias_pendientes:
        jugador_id = a.get("jugador_id")
        fecha = a.get("fecha")

        cursor.execute("SELECT 1 FROM jugadores WHERE id = ?", (jugador_id,))
        if not cursor.fetchone():
            ant_errores += 1
            continue

        cursor.execute(
            "SELECT 1 FROM antropometrias WHERE jugador_id = ? AND fecha = ?",
            (jugador_id, fecha)
        )
        if cursor.fetchone():
            ant_omitidas += 1
            continue

        columnas = list(a.keys())
        valores = list(a.values())
        placeholders = ", ".join(["?"] * len(valores))
        sql = f"INSERT INTO antropometrias ({', '.join(columnas)}) VALUES ({placeholders})"
        try:
            cursor.execute(sql, valores)
            ant_insertadas += 1
        except Exception:
            ant_errores += 1

    conn.commit()
    conn.close()

    return {
        "jug_insertados": jug_insertados,
        "jug_sobreescritos": jug_sobreescritos,
        "jug_omitidos": jug_omitidos,
        "ant_insertadas": ant_insertadas,
        "ant_omitidas": ant_omitidas,
        "ant_errores": ant_errores,
    }


def _flash_resultado(res, errores_jug, errores_ant):
    partes = []
    partes.append(
        f"Jugadores: {res['jug_insertados']} insertados, "
        f"{res['jug_sobreescritos']} actualizados, "
        f"{res['jug_omitidos']} sin cambios."
    )
    partes.append(
        f"Antropometrías: {res['ant_insertadas']} insertadas, "
        f"{res['ant_omitidas']} omitidas."
    )
    if res["ant_errores"] > 0:
        partes.append(f"{res['ant_errores']} filas con errores.")

    todos_errores = errores_jug + errores_ant
    tiene_errores = res["ant_errores"] > 0 or todos_errores

    flash(" ".join(partes), "warning" if tiene_errores else "success")

    if todos_errores:
        for e in todos_errores:
            flash(e, "warning")
